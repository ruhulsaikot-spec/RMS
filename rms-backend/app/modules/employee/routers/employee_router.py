from fastapi import APIRouter
from fastapi import Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
import io
import openpyxl
from openpyxl import Workbook

from app.core.database import get_db

from app.modules.employee.schemas.employee_schema import (
    EmployeeCreate,
    EmployeeUpdate,
    EmployeeResponse,
)

from app.modules.employee.services.employee_service import (
    EmployeeService,
)

router = APIRouter(
    prefix="/employees",
    tags=["Employees"],
)


@router.post(
    "/",
    response_model=EmployeeResponse,
)
async def create_employee(
    payload: EmployeeCreate,
    db: AsyncSession = Depends(get_db),
):

    return await EmployeeService.create_employee(
        db,
        payload,
    )


@router.get(
    "/",
    response_model=list[EmployeeResponse],
)
async def list_employees(
    db: AsyncSession = Depends(get_db),
):

    return await EmployeeService.list_employees(
        db
    )


@router.put(
    "/{employee_id}",
    
)
async def update_employee(
    employee_id: str,
    payload: EmployeeUpdate,
    db: AsyncSession = Depends(get_db),
):

    return await EmployeeService.update_employee(
        db,
        employee_id,
        payload,
    )


@router.delete(
    "/{employee_id}",
)
async def delete_employee(
    employee_id: str,
    db: AsyncSession = Depends(get_db),
):
    return await EmployeeService.delete_employee(
        db,
        employee_id,
    )


@router.get("/bulk/template")
async def download_template():
    """Download Excel template for bulk employee upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Template"

    # Headers
    headers = [
        "Employee ID*", "Full Name*", "Email*", "Mobile",
        "Company Code*", "Department Name*", "Designation Name*",
        "Location Name", "Line Manager Employee ID", "Joining Date (YYYY-MM-DD)", "Status (Active/Inactive)",
    ]
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Sample row
    ws.append([
        "EMP001", "John Doe", "john@example.com", "01700000000",
        "WYZ001", "Engineering", "Software Engineer",
        "Dhaka", "EMP000", "2024-01-01", "Active",
    ])

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"},
    )


@router.post("/bulk/upload")
async def bulk_upload_employees(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Bulk upload employees from Excel file."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed.")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    results = {"success": 0, "failed": 0, "errors": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        try:
            employee_id = str(row[0]).strip() if row[0] else None
            full_name = str(row[1]).strip() if row[1] else None
            email = str(row[2]).strip() if row[2] else None
            mobile = str(row[3]).strip() if row[3] else None
            company_code = str(row[4]).strip() if row[4] else None
            department_name = str(row[5]).strip() if row[5] else None
            designation_name = str(row[6]).strip() if row[6] else None
            location_name = str(row[7]).strip() if row[7] else None
            line_manager_emp_id = str(row[8]).strip() if row[8] else None
            # Handle date - could be string or datetime object from Excel
            raw_date = row[9]
            if raw_date is None:
                joining_date = None
            elif hasattr(raw_date, 'strftime'):
                joining_date = raw_date.strftime("%Y-%m-%d")
            else:
                joining_date = str(raw_date).strip()
            raw_status = row[10]
            status = str(raw_status).strip() if raw_status is not None else "Active"
            print(f"Row {row_idx} status raw: '{raw_status}', parsed: '{status}', is_active: {status.lower() == 'active'}")

            if not all([employee_id, full_name, email, company_code, department_name, designation_name]):
                results["failed"] += 1
                results["errors"].append(f"Row {row_idx}: Missing required fields")
                continue

            await EmployeeService.bulk_create_employee(
                db=db,
                employee_id=employee_id,
                full_name=full_name,
                email=email,
                mobile=mobile,
                company_code=company_code,
                department_name=department_name,
                designation_name=designation_name,
                location_name=location_name,
                line_manager_emp_id=line_manager_emp_id,
                joining_date=joining_date,
                is_active=status.lower() == "active",
            )
            results["success"] += 1

        except Exception as e:
            results["failed"] += 1
            results["errors"].append(f"Row {row_idx}: {str(e)}")

    return results