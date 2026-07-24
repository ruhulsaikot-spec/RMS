from fastapi import APIRouter
from fastapi import Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.auth.dependencies import require_permission

from app.core.database import get_db
from fastapi import Query
from app.core.schemas.api_response import ApiResponse
from app.modules.user.schemas.user_schema import (
    AssignRoleRequest,
)

from app.modules.user.schemas.user_schema import (
    UserCreate,
    UserResponse,
    UserDetailResponse,
    UserUpdate,
    UserListResponse,
    ResetPasswordRequest,
)

from app.modules.user.services.user_service import (
    UserService,
)

router = APIRouter(
    prefix="/users",
    tags=["Users"],
)


@router.post(
    "/",
    response_model=UserDetailResponse,
)
async def create_user(
    payload: UserCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(
        require_permission(
            "user:create"
        )
    ),
):
    return await UserService.create_user(
    db,
    payload,
    current_user["id"],
)


@router.get(
    "/{user_id}",
    response_model=UserDetailResponse,
)
async def get_user_by_id(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(
        require_permission(
            "user:read"
        )
    ),
):
    user = await UserService.get_user_by_id(
        db,
        user_id,
    )

    return UserDetailResponse.model_validate(
        user
    )    

@router.put(
    "/{user_id}",
    response_model=UserDetailResponse,
)
async def update_user(
    user_id: str,
    payload: UserUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(
        require_permission(
            "user:update"
        )
    ),
):
    return await UserService.update_user(
        db,
        user_id,
        payload,
        current_user["id"],
    )

@router.get("/")
async def get_users(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    search: str | None = None,
    is_active: bool | None = None,
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(
        require_permission(
            "user:list"
        )
    ),
):
    result = await UserService.get_users(
        db,
        page,
        size,
        search,
        is_active,
        sort_by,
        sort_order,
    )

    return ApiResponse(
        message="Users fetched successfully",
        data=[
            UserResponse.model_validate(user)
            for user in result["items"]
        ],
        meta={
            "total": result["total"],
            "page": result["page"],
            "size": result["size"],
        },
    )
@router.delete(
    "/{user_id}",
)
async def delete_user(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(
        require_permission(
            "user:delete"
        )
    ),
):
    return await UserService.delete_user(
    db,
    user_id,
    current_user["id"],
    )

@router.post(
    "/{user_id}/roles",
)
async def assign_roles(
    user_id: str,
    payload: AssignRoleRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(
        require_permission(
            "user:manage_roles"
        )
    ),
):
    return await UserService.assign_roles(
        db,
        user_id,
        payload.role_ids,
    )


@router.post(
    "/{user_id}/reset-password",
)
async def reset_password(
    user_id: str,
    payload: ResetPasswordRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(
        require_permission(
            "user:update"
        )
    ),
):
    return await UserService.reset_password(
        db,
        user_id,
        payload.password,
        current_user["id"],
    )


@router.get("/bulk/template")
async def download_user_template():
    """Download Excel template for bulk user upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "User Template"

    headers = [
        "Employee ID*", "Full Name*", "Email*", "Phone",
        "Department Name*", "Designation Name*",
        "Role Name*", "Password*", "Status (Active/Inactive)",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 24

    ws.append([
        "EMP001", "John Doe", "john@example.com", "01700000000",
        "Engineering", "Software Engineer",
        "Employee", "Password@123", "Active",
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_template.xlsx"},
    )


@router.post("/bulk/upload")
async def bulk_upload_users(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_permission("user:create")),
):
    """Bulk upload users from Excel file."""
    from sqlalchemy import text

    if not file.filename.endswith((".xlsx", ".xls")):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Only Excel files are allowed.")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    results = {"success": 0, "failed": 0, "errors": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        try:
            employee_id = str(row[0]).strip() if row[0] else None
            full_name = str(row[1]).strip() if row[1] else None
            email = str(row[2]).strip() if row[2] else None
            phone = str(row[3]).strip() if row[3] else None
            department_name = str(row[4]).strip() if row[4] else None
            designation_name = str(row[5]).strip() if row[5] else None
            role_names_raw = str(row[6]).strip() if row[6] else ""
            role_names = [r.strip() for r in role_names_raw.split(",") if r.strip()]
            password = str(row[7]).strip() if row[7] else None
            status = str(row[8]).strip() if row[8] else "Active"

            if not all([employee_id, full_name, email, department_name, designation_name, role_names_raw, password]):
                results["failed"] += 1
                results["errors"].append(f"Row {row_idx}: Missing required fields")
                continue

            # Lookup department
            dept_result = await db.execute(
                text("SELECT id FROM departments WHERE LOWER(name) = LOWER(:name) LIMIT 1"),
                {"name": department_name}
            )
            dept_row = dept_result.fetchone()
            department_id = dept_row[0] if dept_row else None

            # Lookup designation
            desig_result = await db.execute(
                text("SELECT id FROM designations WHERE LOWER(name) = LOWER(:name) LIMIT 1"),
                {"name": designation_name}
            )
            desig_row = desig_result.fetchone()
            designation_id = desig_row[0] if desig_row else None

            # Lookup roles by name or display_name
            role_ids = []
            for rname in role_names:
                role_result = await db.execute(
                    text("""
                        SELECT id FROM roles 
                        WHERE LOWER(name) = LOWER(:name) 
                        OR LOWER(display_name) = LOWER(:name)
                        LIMIT 1
                    """),
                    {"name": rname}
                )
                role_row = role_result.fetchone()
                if role_row:
                    role_ids.append(str(role_row[0]))
                else:
                    results["errors"].append(f"Row {row_idx}: Role '{rname}' not found, skipping role")

            if not role_ids:
                results["failed"] += 1
                results["errors"].append(f"Row {row_idx}: No valid roles found")
                continue

            # Create user
            payload = UserCreate(
                employee_id=employee_id,
                full_name=full_name,
                email=email,
                phone=phone,
                department_id=department_id,
                designation_id=designation_id,
                password=password,
                is_active=status.lower() == "active",
            )

            user = await UserService.create_user(db, payload, current_user["id"])

            # Assign roles
            await UserService.assign_roles(db, user.id, role_ids)

            results["success"] += 1

        except Exception as e:
            results["failed"] += 1
            results["errors"].append(f"Row {row_idx}: {str(e)}")

    return results